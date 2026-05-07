"""
CMスケ表 - PDF→Excel変換 APIエンドポイント
Vercel Python Serverless Function (Flask)
"""

from flask import Flask, request, send_file, jsonify
import pdfplumber
from openpyxl import load_workbook
import io
import re
import os
import tempfile
from datetime import time, datetime, timedelta

app = Flask(__name__)

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "template.xlsx")
DOW_JA = "月火水木金土日"
DOW_TO_COL = {0: 1, 1: 5, 2: 9, 3: 13, 4: 17, 5: 21, 6: 25}


# ── ユーティリティ ──────────────────────────────────

def zen2han(s: str) -> str:
    return s.translate(str.maketrans(
        "ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ０１２３４５６７８９",
        "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789"
    ))


def _detect_format(text: str, tables: list) -> str:
    """PDFフォーマット自動判定: 'fct_tuf' または 'ftv'"""
    if "契約コード" in text:
        return "fct_tuf"
    if "契約番号" in text:
        return "ftv"
    # テーブルセルのヘッダーパターンで判定
    for table in tables:
        for row in table:
            for cell in (row or []):
                s = str(cell or "")
                if re.search(r"≪\d{2}/\d{2}", s):
                    return "fct_tuf"
                if re.search(r"\d{2}/\d{2}\([月火水木金土日]\)", s):
                    return "ftv"
    return "fct_tuf"  # デフォルト


# ── FCT/TUF フォーマット ──────────────────────────────

def _parse_meta_fct_tuf(text: str) -> dict:
    def ex(pattern, default=""):
        m = re.search(pattern, text)
        return m.group(1).strip() if m else default

    meta = {
        "format":        "fct_tuf",
        "contract_code": zen2han(ex(r"契約コード\s*：\s*([A-Za-zＡ-Ｚ０-９0-9]+")),
        "sponsor":       ex(r"スポンサー\s*：\s*(.+?)\s+A単価"),
        "agency":        ex(r"広告会社\s*：\s*(.+?)\s+評価欄帯域"),
        "person":        ex(r"外勤\s*：\s*(.+)"),
        "product":       ex(r"商品名\s*：\s*(.+?)\s+枠取りパターン"),
        "seconds":       int(ex(r"枠取り秒数\s*：\s*(\d+)", "15")),
    }

    sm = re.search(r"^([^\s/]+／[^\s]+)\s+\d+／\d+", text, re.MULTILINE)
    meta["station"] = sm.group(1).split("／")[0] if sm else ""

    pm = re.search(r"契約期間\s*：\s*(\d{4}/\d{2}/\d{2})～(\d{4}/\d{2}/\d{2})", text)
    if not pm:
        raise ValueError("契約期間が取得できません")
    meta["period_start"] = datetime.strptime(pm.group(1), "%Y/%m/%d")
    meta["period_end"]   = datetime.strptime(pm.group(2), "%Y/%m/%d")

    tm = re.search(r"枠取りパターン[^\n]+?(\d+)本\s*$", text, re.MULTILINE)
    meta["total_count"] = int(tm.group(1)) if tm else 0
    return meta


def _parse_schedule_fct_tuf(tables: list) -> dict:
    schedule = {}
    day_re   = re.compile(r"≪(\d{2}/\d{2})\s*[月火水木金土日]≫")
    entry_re = re.compile(r"ﾚ\s*(\d{4})([TS])\s*([ABCS])\s*(\d+)\"")

    for table in tables:
        current_days = {}
        for row in table:
            if not row:
                continue
            day_cols = {
                ci: day_re.search(str(cell)).group(1)
                for ci, cell in enumerate(row)
                if cell and day_re.search(str(cell))
            }
            if day_cols:
                current_days = day_cols
                for d in current_days.values():
                    schedule.setdefault(d, [])
                continue
            for ci, cell in enumerate(row):
                if not cell or ci not in current_days:
                    continue
                em = entry_re.search(str(cell))
                if not em:
                    continue
                raw = int(em.group(1))
                past = raw >= 2400
                if past:
                    raw -= 2400
                h, m = raw // 100, raw % 100
                schedule[current_days[ci]].append({
                    "time_obj":      time(h, m),
                    "sb_pt":         "PT" if em.group(2) == "T" else "SB",
                    "seconds":       int(em.group(4)),
                    "past_midnight": past,
                })
    return schedule


# ── FTV フォーマット ──────────────────────────────────

def _parse_meta_ftv(text: str) -> dict:
    def ex(pattern, default=""):
        m = re.search(pattern, text)
        return m.group(1).strip() if m else default

    meta = {"format": "ftv"}

    # 契約番号
    meta["contract_code"] = zen2han(ex(
        r"契約番号\s*[：:]\s*([A-Za-zＡ-Ｚ０-９0-9]+)"
    ))

    # スポンサー（「スポンサー [名前]」形式 or 「スポンサー：[名前]」形式）
    m_sp = re.search(
        r"スポンサー\s*[：:]?\s*(.+?)\s+(?:商品名称|商品名|A単価)", text
    )
    meta["sponsor"] = m_sp.group(1).strip() if m_sp else ex(r"スポンサー\s+(\S+)")

    # 代理店
    m_ag = re.search(r"代理店\s*[：:]?\s*(.+?)\s+(?:契約期間|外勤)", text)
    meta["agency"] = m_ag.group(1).strip() if m_ag else ""

    # 商品名
    meta["product"] = ex(r"商品名称?\s*[：:]?\s*(.+?)\s+(?:代理店|秒数|枠)")

    # 担当者
    meta["person"] = ex(r"外勤\s*[：:]?\s*(.+)")

    # 秒数
    meta["seconds"] = int(ex(r"(?:枠取り)?秒数\s*[：:]?\s*(\d+)", "15"))

    # 放送局（FTV 福島テレビ 等の局名行）
    m_st = re.search(r"(FTV|TUF|FCT|KFB|福島[^\s]*テレビ[^\s]*)\s+SPOT", text)
    meta["station"] = m_st.group(1) if m_st else "FTV"

    # 契約期間: 「2024年 2月10日 ～ 2024年 3月 2日」
    pm = re.search(
        r"(\d{4})年\s*(\d+)月\s*(\d+)日\s*[〜～]\s*(\d{4})年\s*(\d+)月\s*(\d+)日",
        text
    )
    if pm:
        meta["period_start"] = datetime(int(pm.group(1)), int(pm.group(2)), int(pm.group(3)))
        meta["period_end"]   = datetime(int(pm.group(4)), int(pm.group(5)), int(pm.group(6)))
    else:
        # 別パターン: YYYY/MM/DD
        pm2 = re.search(r"(\d{4}/\d{2}/\d{2})[〜～](\d{4}/\d{2}/\d{2})", text)
        if pm2:
            meta["period_start"] = datetime.strptime(pm2.group(1), "%Y/%m/%d")
            meta["period_end"]   = datetime.strptime(pm2.group(2), "%Y/%m/%d")
        else:
            raise ValueError("契約期間が取得できません")

    # 総本数
    tm = re.search(r"(\d+)\s*本", text)
    meta["total_count"] = int(tm.group(1)) if tm else 0
    return meta


def _parse_schedule_ftv(tables: list) -> dict:
    """FTVスケジュール解析: 'MM/DD(曜)' ヘッダー + 'H:MM[P] SS RANK' エントリ"""
    schedule = {}
    day_re   = re.compile(r"(\d{2}/\d{2})\([月火水木金土日]\)")
    # H:MM[P] SS RANK — Pあり=PT、なし=SB
    entry_re = re.compile(r"(\d+):(\d{2})(P?)\s+(\d+)\s+([ABCS])")

    for table in tables:
        current_days = {}
        for row in table:
            if not row:
                continue
            cells = [str(c) if c is not None else "" for c in row]

            # 日付ヘッダー行の検出（全セルの半数以上がMM/DD(曜)形式）
            day_matches = {
                ci: day_re.search(c).group(1)
                for ci, c in enumerate(cells)
                if day_re.search(c)
            }
            if len(day_matches) >= 3:
                current_days = day_matches
                for d in current_days.values():
                    schedule.setdefault(d, [])
                continue

            # エントリ行
            for ci, cell in enumerate(cells):
                if not cell.strip() or ci not in current_days:
                    continue
                day_key = current_days[ci]
                for line in cell.split("\n"):
                    line = line.strip()
                    em = entry_re.search(line)
                    if not em:
                        continue
                    h    = int(em.group(1))
                    m_   = int(em.group(2))
                    past = h >= 24
                    if past:
                        h -= 24
                    schedule[day_key].append({
                        "time_obj":      time(h, m_),
                        "sb_pt":         "PT" if em.group(3) == "P" else "SB",
                        "seconds":       int(em.group(4)),
                        "past_midnight": past,
                    })

    return schedule


# ── メイン解析エントリ ────────────────────────────────

def parse_meisai(pdf_bytes: bytes) -> dict:
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        full_text = "\n".join(p.extract_text() or "" for p in pdf.pages)
        all_tables = [t for p in pdf.pages for t in p.extract_tables()]

    fmt = _detect_format(full_text, all_tables)

    if fmt == "ftv":
        meta = _parse_meta_ftv(full_text)
        meta["schedule"] = _parse_schedule_ftv(all_tables)
    else:
        meta = _parse_meta_fct_tuf(full_text)
        meta["schedule"] = _parse_schedule_fct_tuf(all_tables)

    return meta


# ── Excel生成 ─────────────────────────────────────

def _excel_time(entry):
    t = entry["time_obj"]
    if entry["past_midnight"]:
        return datetime(1900, 1, 1, t.hour, t.minute)
    return t


def _cm_abbr(date_str: str, year: int, materials: list) -> str:
    dt = datetime(year, int(date_str[:2]), int(date_str[3:]))
    for mat in materials:
        pm = re.match(r"(\d+)/(\d+)[〜~](\d+)/(\d+)", mat.get("period", ""))
        if pm:
            if datetime(year, int(pm.group(1)), int(pm.group(2))) <= dt <= \
               datetime(year, int(pm.group(3)), int(pm.group(4))):
                return mat["abbr"]
    return materials[0]["abbr"] if materials else ""


def build_excel(meta: dict, materials: list, station_name: str,
                station_person: str, doc_date: str) -> bytes:
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active
    year = meta["period_start"].year

    if doc_date:
        try:
            ws.cell(2, 22, datetime.strptime(doc_date, "%Y/%m/%d"))
        except ValueError:
            pass

    loc = station_name or meta.get("station", "")
    if station_person:
        loc += f"　{station_person}様"
    ws.cell(3, 2, loc)

    ws.cell(4, 2, meta["sponsor"])
    ws.cell(4, 10, f"契約No.{meta['contract_code']}")

    ps, pe = meta["period_start"], meta["period_end"]
    ws.cell(5, 2, (
        f"{ps.year-2000}年{ps.month}月{ps.day}日({DOW_JA[ps.weekday()]})"
        f"～{pe.month}月{pe.day}日({DOW_JA[pe.weekday()]})"
    ))
    ws.cell(5, 18, meta["total_count"])

    for i, mat in enumerate(materials):
        r = 8 + i
        ws.cell(r, 1,  meta["sponsor"])
        ws.cell(r, 3,  mat.get("name", ""))
        ws.cell(r, 9,  int(mat.get("seconds", meta["seconds"])))
        ws.cell(r, 10, mat.get("abbr", ""))
        ws.cell(r, 12, mat.get("material", "OL"))
        ws.cell(r, 14, mat.get("status", "送"))
        ws.cell(r, 16, mat.get("period", ""))

    schedule = meta["schedule"]
    if not schedule:
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    date_objs = {d: datetime(year, int(d[:2]), int(d[3:])) for d in schedule}
    sorted_d  = sorted(date_objs, key=lambda d: date_objs[d])
    first_dt, last_dt = date_objs[sorted_d[0]], date_objs[sorted_d[-1]]
    first_mon = first_dt - timedelta(days=first_dt.weekday())

    weeks, cur = [], first_mon
    while cur <= last_dt:
        week = []
        for i in range(7):
            d = cur + timedelta(days=i)
            ds = f"{d.month:02d}/{d.day:02d}"
            week.append((ds, d, schedule.get(ds, [])))
        if any(e for _, _, e in week):
            weeks.append(week)
        cur += timedelta(days=7)

    crow = 13
    for week in weeks:
        for ds, dt, entries in week:
            if entries:
                ws.cell(crow, DOW_TO_COL[dt.weekday()],
                        f"{dt.month}/{dt.day}({DOW_JA[dt.weekday()]})")
        mx = max(len(e) for _, _, e in week)
        for idx in range(mx):
            row = crow + 1 + idx
            for ds, dt, entries in week:
                if idx >= len(entries):
                    continue
                e   = entries[idx]
                col = DOW_TO_COL[dt.weekday()]
                ws.cell(row, col,     _excel_time(e))
                ws.cell(row, col + 1, e["sb_pt"])
                ws.cell(row, col + 2, e["seconds"])
                ws.cell(row, col + 3, _cm_abbr(ds, year, materials))
        crow += 1 + mx + 1

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── Flask ルート ──────────────────────────────────

@app.route("/api/generate", methods=["POST"])
def generate():
    meisai_file = request.files.get("meisai")
    if not meisai_file:
        return jsonify({"error": "明細リストPDFが必要です"}), 400

    try:
        meta = parse_meisai(meisai_file.read())
    except Exception as e:
        return jsonify({"error": f"PDF解析エラー: {str(e)}"}), 422

    import json as _json
    materials_raw = request.form.get("materials", "[]")
    try:
        materials = _json.loads(materials_raw)
    except Exception:
        materials = []

    excel_bytes = build_excel(
        meta=meta,
        materials=materials,
        station_name=request.form.get("station", ""),
        station_person=request.form.get("person", ""),
        doc_date=request.form.get("date", ""),
    )

    sponsor_short = re.sub(r"[^\w]", "", meta.get("sponsor", "output"))[:20]
    filename = f"CMスケ表_{sponsor_short}.xlsx"

    return send_file(
        io.BytesIO(excel_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/api/health")
def health():
    return jsonify({"status": "ok"})
